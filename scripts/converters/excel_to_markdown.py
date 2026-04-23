from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from converters.common import (
    FileReport,
    find_header_column,
    image_resource_markdown,
    import_or_fail,
    make_image_record,
    normalize_cell,
    row_value,
    slugify_filename,
    table_to_markdown,
    write_md,
    yaml_frontmatter,
)


def parse_anchor_row(anchor: str) -> int | None:
    match = re.search(r"(\d+)$", str(anchor or "").strip())
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def find_semantic_header_index(rows: list[list[Any]]) -> int:
    header_keywords = ("产品名称", "产品类别", "产品原料", "产品简介", "产品功效", "体质", "商品名称", "图片")
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows):
        row_text = " ".join(normalize_cell(cell) for cell in row)
        score = sum(1 for keyword in header_keywords if keyword in row_text)
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def product_fields(header: list[Any], row: list[Any]) -> dict[str, str]:
    name_idx = find_header_column(header, "产品", "名称")
    category_idx = find_header_column(header, "产品", "类别")
    material_idx = find_header_column(header, "产品", "原料")
    intro_idx = find_header_column(header, "产品", "简介")
    effect_idx = find_header_column(header, "产品", "功效")
    constitution_idx = find_header_column(header, "体质")

    fields = {
        "产品名称": row_value(row, name_idx),
        "产品类别": row_value(row, category_idx),
        "适用体质": row_value(row, constitution_idx),
        "产品原料": row_value(row, material_idx),
        "产品简介": row_value(row, intro_idx),
        "产品功效": row_value(row, effect_idx),
    }
    if not fields["产品名称"]:
        non_empty = [normalize_cell(cell) for cell in row if normalize_cell(cell)]
        fields["产品名称"] = non_empty[0] if non_empty else "未命名商品"
    return fields


def semantic_image_caption(source_file: Path, sheet_name: str, fields: dict[str, str], image_index: int) -> str:
    product_name = fields.get("产品名称") or ""
    category = fields.get("产品类别") or sheet_name
    constitution = fields.get("适用体质") or ""
    pieces = []
    if product_name:
        pieces.append(f"{product_name}产品图")
    else:
        pieces.append(f"{source_file.stem}{sheet_name}图片{image_index}")
    if category:
        pieces.append(category)
    if constitution:
        pieces.append(f"{constitution}适用")
    return "，".join(pieces)


def associate_excel_images(
    image_records: list[dict],
    source_rows: list[int],
    semantic_header_index: int,
) -> dict[int, list[dict]]:
    data_source_rows = source_rows[semantic_header_index + 1 :]
    if not image_records or not data_source_rows:
        return {}

    anchor_rows = [parse_anchor_row(record.get("anchor")) for record in image_records]
    exact_count = sum(1 for row in anchor_rows if row in data_source_rows)
    shifted_count = 0
    offset = 0
    if anchor_rows[0] is not None:
        offset = anchor_rows[0] - data_source_rows[0]
        shifted_count = sum(1 for row in anchor_rows if row is not None and row - offset in data_source_rows)

    use_shifted = shifted_count > exact_count
    associated: dict[int, list[dict]] = {}
    for index, record in enumerate(image_records):
        anchor_row = anchor_rows[index]
        target_row = None
        if anchor_row is not None:
            candidate = anchor_row - offset if use_shifted else anchor_row
            if candidate in data_source_rows:
                target_row = candidate
        if target_row is None and index < len(data_source_rows):
            target_row = data_source_rows[index]
        if target_row is None:
            continue
        record["associated_source_row"] = target_row
        associated.setdefault(target_row, []).append(record)
    return associated


def product_resource_markdown(
    source_file: Path,
    sheet_name: str,
    header: list[Any],
    product_rows: list[tuple[int, list[Any]]],
    images_by_source_row: dict[int, list[dict]],
) -> str:
    if not product_rows:
        return ""

    lines = ["## 商品资料（含图片）", ""]
    for source_row, row in product_rows:
        fields = product_fields(header, row)
        product_name = fields.get("产品名称") or f"商品行 {source_row}"
        lines.extend([f"### {product_name}", ""])
        for label in ("产品类别", "适用体质", "产品原料", "产品简介", "产品功效"):
            value = fields.get(label)
            if value:
                lines.append(f"{label}: {value}")

        image_records = images_by_source_row.get(source_row, [])
        if image_records:
            lines.append("")
            lines.append("商品图片:")
            for image_index, record in enumerate(image_records, start=1):
                caption = semantic_image_caption(source_file, sheet_name, fields, image_index)
                record["caption"] = caption
                record["associated_product_name"] = product_name
                record["associated_constitution"] = fields.get("适用体质", "")
                lines.extend(
                    [
                        f"image_id: {record['image_id']}",
                        f"image_caption: {caption}",
                        f"image_url: {record['image_url']}",
                        f"![{caption}](<{record['image_url']}>)",
                    ]
                )
        lines.append("")
    return "\n".join(lines)


def is_empty_row(row: list[Any]) -> bool:
    return all(normalize_cell(cell) == "" for cell in row)


def skip_sheet_reason(ws, rows: list[list[Any]]) -> str:
    sheet_state = str(getattr(ws, "sheet_state", "visible") or "visible").lower()
    sheet_name = str(getattr(ws, "title", "") or "")
    if sheet_state != "visible":
        return f"hidden sheet: {sheet_state}"
    if sheet_name.startswith("WpsReserved_"):
        return "internal WPS reserved sheet"
    if not rows:
        return "empty sheet or no effective cell data"
    if len(rows) == 1:
        return "only one row, no effective data rows"
    return ""


def mark_sheet_skipped(report: FileReport, ws, reason: str) -> None:
    report.skipped_sheets.append({"sheet_name": ws.title, "reason": reason})


def fill_merged_cells(ws, rows: list[list[Any]]) -> None:
    if not rows:
        return
    for merged_range in ws.merged_cells.ranges:
        top_row = merged_range.min_row - 1
        left_col = merged_range.min_col - 1
        if top_row >= len(rows) or left_col >= len(rows[0]):
            continue
        value = rows[top_row][left_col]
        for row_idx in range(merged_range.min_row - 1, min(merged_range.max_row, len(rows))):
            for col_idx in range(merged_range.min_col - 1, min(merged_range.max_col, len(rows[0]))):
                rows[row_idx][col_idx] = value


def extract_sheet_rows(ws) -> tuple[list[int], list[list[Any]]]:
    raw_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
    fill_merged_cells(ws, raw_rows)
    row_indexes = [idx for idx, row in enumerate(raw_rows) if not is_empty_row(row)]
    if not row_indexes:
        return [], []
    col_indexes = [
        col_idx
        for col_idx in range(len(raw_rows[0]))
        if any(normalize_cell(raw_rows[row_idx][col_idx]) for row_idx in row_indexes)
    ]
    return (
        [idx + 1 for idx in row_indexes],
        [[raw_rows[row_idx][col_idx] for col_idx in col_indexes] for row_idx in row_indexes],
    )


def excel_anchor(image) -> str:
    marker = getattr(getattr(image, "anchor", None), "_from", None)
    if marker is None:
        return "unknown"
    try:
        from openpyxl.utils import get_column_letter

        return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"
    except Exception:
        return "unknown"


def save_excel_images(ws, source_file: Path, assets_dir: Path, asset_base_url: str) -> list[dict]:
    records = []
    for index, image in enumerate(getattr(ws, "_images", []), start=1):
        ext = (getattr(image, "format", None) or "png").lower().lstrip(".")
        if ext == "jpeg":
            ext = "jpg"
        stem = f"{slugify_filename(source_file.stem)}__{slugify_filename(ws.title)}__image{index:03d}"
        filename = f"{stem}.{ext}"
        relative_path = f"assets/{filename}"
        (assets_dir / filename).write_bytes(image._data())
        image_id = f"{slugify_filename(source_file.stem, 40)}_{slugify_filename(ws.title, 30)}_{index:03d}"
        records.append(
            make_image_record(
                source_file=source_file,
                source_type="excel",
                relative_path=relative_path,
                asset_base_url=asset_base_url,
                image_id=image_id,
                caption=f"{source_file.stem} {ws.title} 图片 {index}",
                sheet_name=ws.title,
                anchor=excel_anchor(image),
            )
        )
    return records


def convert_excel(source_file: Path, output_dir: Path, assets_dir: Path, rows_per_file: int, generated_at: str, asset_base_url: str, image_index: list[dict]) -> FileReport:
    openpyxl = import_or_fail("openpyxl", "openpyxl")
    report = FileReport(source_file.name, "excel")
    workbook = openpyxl.load_workbook(source_file, data_only=True)
    report.sheets = len(workbook.worksheets)

    for ws in workbook.worksheets:
        source_rows, rows = extract_sheet_rows(ws)
        skip_reason = skip_sheet_reason(ws, rows)
        if skip_reason:
            mark_sheet_skipped(report, ws, skip_reason)
            continue

        image_records = save_excel_images(ws, source_file, assets_dir, asset_base_url)
        report.images += len(image_records)
        image_index.extend(image_records)

        if not rows:
            rows = [["说明"], ["该工作表为空或未识别到有效数据。"]]
            source_rows = [1, 1]

        semantic_header_index = find_semantic_header_index(rows)
        semantic_header = rows[semantic_header_index] if rows else []
        semantic_header_source_row = source_rows[semantic_header_index] if semantic_header_index < len(source_rows) else source_rows[0]
        rows_by_source_row = dict(zip(source_rows, rows))
        images_by_source_row = associate_excel_images(image_records, source_rows, semantic_header_index)
        associated_images = {
            id(record)
            for records in images_by_source_row.values()
            for record in records
        }
        unmatched_images = [record for record in image_records if id(record) not in associated_images]

        header = rows[0]
        data_rows = rows[1:] or []
        chunks = []
        if data_rows:
            for start in range(0, len(data_rows), rows_per_file):
                chunks.append((source_rows[start + 1 : start + 1 + rows_per_file], [header] + data_rows[start : start + rows_per_file]))
        else:
            chunks.append((source_rows, rows))

        for part_index, (chunk_source_rows, chunk_rows) in enumerate(chunks, start=1):
            row_range = "empty" if not chunk_source_rows else f"{chunk_source_rows[0]}-{chunk_source_rows[-1]}"
            output_name = f"{slugify_filename(source_file.stem)}__{slugify_filename(ws.title)}__part{part_index:03d}.md"
            output_path = output_dir / output_name
            chunk_product_rows = [
                (source_row, rows_by_source_row[source_row])
                for source_row in chunk_source_rows
                if source_row > semantic_header_source_row and source_row in rows_by_source_row
            ]
            chunk_image_ids = {
                id(record)
                for source_row, _ in chunk_product_rows
                for record in images_by_source_row.get(source_row, [])
            }
            part_images = [
                record
                for record in image_records
                if id(record) in chunk_image_ids
            ]
            if part_index == len(chunks):
                part_images.extend(unmatched_images)
            for record in part_images:
                record["markdown_file"] = output_name
            unassociated_resource = image_resource_markdown(unmatched_images) if part_index == len(chunks) else ""
            write_md(
                output_path,
                [
                    yaml_frontmatter(
                        {
                            "source_file": source_file.name,
                            "source_type": "excel",
                            "sheet_name": ws.title,
                            "row_range": row_range,
                            "generated_at": generated_at,
                        }
                    ),
                    "",
                    f"# {source_file.stem}",
                    "",
                    f"## 工作表：{ws.title}",
                    "",
                    f"### 数据行范围：{row_range}",
                    "",
                    table_to_markdown(chunk_rows),
                    "",
                    product_resource_markdown(
                        source_file=source_file,
                        sheet_name=ws.title,
                        header=semantic_header,
                        product_rows=chunk_product_rows,
                        images_by_source_row=images_by_source_row,
                    ),
                    "",
                    unassociated_resource,
                ],
            )
            report.markdown_files.append(output_name)

    workbook.close()
    return report
